"""
excel_exporter.py
------------------
Converts structured sheet data from the LLM into a properly formatted .xlsx file.
Follows financial document conventions:
- Blue for hardcoded inputs
- Bold for headers and totals
- Currency formatting
- Alternating row colors
- Per-sheet tabs
"""

from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter


# ── Color constants ────────────────────────────────────────────────────────────
COLOR_HEADER_BG    = "1F3864"   # Dark navy — sheet header row background
COLOR_HEADER_FONT  = "FFFFFF"   # White text on header
COLOR_TOTAL_BG     = "D6E4F0"   # Light blue — totals rows
COLOR_TOTAL_FONT   = "1F3864"   # Navy text on totals
COLOR_SECTION_BG   = "EBF3FB"   # Very light blue — section header rows
COLOR_SECTION_FONT = "1F3864"
COLOR_INPUT        = "0000FF"   # Blue — hardcoded input values (financial convention)
COLOR_ALT_ROW      = "F5F8FF"   # Alternating row tint
COLOR_BORDER       = "BDD7EE"   # Light border color


def _make_border(style="thin"):
    side = Side(style=style, color=COLOR_BORDER)
    return Border(left=side, right=side, top=side, bottom=side)


def _make_header_fill():
    return PatternFill("solid", fgColor=COLOR_HEADER_BG)

def _make_total_fill():
    return PatternFill("solid", fgColor=COLOR_TOTAL_BG)

def _make_section_fill():
    return PatternFill("solid", fgColor=COLOR_SECTION_BG)

def _make_alt_fill():
    return PatternFill("solid", fgColor=COLOR_ALT_ROW)


def _is_numeric(value: str) -> bool:
    """Check if a string value looks like a number."""
    try:
        cleaned = str(value).replace(",", "").replace("₹", "").replace("$", "").replace("%", "").strip()
        if cleaned in ("", "-", "–"):
            return False
        float(cleaned)
        return True
    except (ValueError, AttributeError):
        return False


def _parse_number(value: str):
    """Parse a numeric string to float."""
    try:
        cleaned = str(value).replace(",", "").replace("₹", "").replace("$", "").strip()
        return float(cleaned)
    except (ValueError, AttributeError):
        return value


def write_sheet(wb: Workbook, sheet_data: dict):
    """Write one sheet of tabular data to the workbook."""
    # Excel forbids: / \ ? * [ ] : in sheet names, and max 31 chars
    _invalid = r'/\?*[]:'
    raw_name = sheet_data.get("sheet_name", "Sheet")
    sheet_name = "".join(c if c not in _invalid else "-" for c in raw_name)[:31].strip()
    if not sheet_name:
        sheet_name = "Sheet"
    headers    = sheet_data.get("headers", [])
    rows       = sheet_data.get("rows", [])
    header_rows = set(sheet_data.get("header_rows", []))
    totals_rows = set(sheet_data.get("totals_rows", []))
    notes       = sheet_data.get("notes", "")
    description = sheet_data.get("description", "")

    ws = wb.create_sheet(title=sheet_name)

    # ── Sheet title ────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 1))
    title_cell = ws.cell(row=1, column=1, value=sheet_name.upper())
    title_cell.font = Font(name="Arial", bold=True, size=13, color=COLOR_HEADER_FONT)
    title_cell.fill = _make_header_fill()
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Description row (if present) ──────────────────────────────────────────
    start_row = 2
    if description:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
        desc_cell = ws.cell(row=2, column=1, value=description)
        desc_cell.font = Font(name="Arial", italic=True, size=9, color="666666")
        desc_cell.alignment = Alignment(horizontal="left", vertical="center")
        start_row = 3

    # ── Column headers row ─────────────────────────────────────────────────────
    header_row_num = start_row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row_num, column=col_idx, value=header)
        cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_HEADER_FONT)
        cell.fill = _make_header_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _make_border()
    ws.row_dimensions[header_row_num].height = 22

    # ── Data rows ──────────────────────────────────────────────────────────────
    data_start = header_row_num + 1
    for row_idx, row in enumerate(rows):
        excel_row = data_start + row_idx
        is_section_header = row_idx in header_rows
        is_total          = row_idx in totals_rows
        is_alt            = (row_idx % 2 == 0) and not is_section_header and not is_total

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=excel_row, column=col_idx)

            # Try to store as number if numeric (except first column — usually labels)
            if col_idx > 1 and _is_numeric(value):
                cell.value = _parse_number(value)
                cell.number_format = '#,##0.00'
                # Blue font for hardcoded inputs (financial convention)
                base_color = COLOR_INPUT if not is_total else COLOR_TOTAL_FONT
            else:
                cell.value = str(value) if value is not None else ""
                base_color = "000000"

            # Styling based on row type
            if is_section_header:
                cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_SECTION_FONT)
                cell.fill = _make_section_fill()
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center", vertical="center")
            elif is_total:
                cell.font = Font(name="Arial", bold=True, size=10, color=COLOR_TOTAL_FONT)
                cell.fill = _make_total_fill()
                cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "right", vertical="center")
            else:
                cell.font = Font(name="Arial", size=10, color=base_color)
                if is_alt:
                    cell.fill = _make_alt_fill()
                cell.alignment = Alignment(
                    horizontal="left" if col_idx == 1 else "right",
                    vertical="center",
                    indent=1 if col_idx == 1 and not is_section_header else 0
                )

            cell.border = _make_border()

        ws.row_dimensions[excel_row].height = 18

    # ── Notes row ─────────────────────────────────────────────────────────────
    if notes:
        notes_row = data_start + len(rows) + 1
        ws.merge_cells(
            start_row=notes_row, start_column=1,
            end_row=notes_row, end_column=max(len(headers), 1)
        )
        notes_cell = ws.cell(row=notes_row, column=1, value=f"Notes: {notes}")
        notes_cell.font = Font(name="Arial", italic=True, size=9, color="888888")
        notes_cell.alignment = Alignment(horizontal="left", wrap_text=True)

    # ── Auto-size columns ─────────────────────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(headers[col_idx - 1])) if col_idx <= len(headers) else 10
        for row in rows:
            if col_idx <= len(row):
                max_len = max(max_len, len(str(row[col_idx - 1])))
        # First column wider (labels), others standard
        if col_idx == 1:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
        else:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 20)

    # Freeze top rows (title + headers)
    ws.freeze_panes = ws.cell(row=data_start, column=1)


def generate_excel_file(excel_data: dict) -> BytesIO:
    """
    Main entry point. Takes the structured excel_data dict from the LLM
    and produces a formatted .xlsx BytesIO stream.

    excel_data format:
    {
      "title": "...",
      "sheets": [ { sheet_data }, ... ]
    }
    """
    wb = Workbook()

    # Remove default empty sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheets = excel_data.get("sheets", [])

    # ── Cover sheet ───────────────────────────────────────────────────────────
    cover = wb.create_sheet(title="Cover", index=0)
    title = excel_data.get("title", "Financial Document")

    cover.merge_cells("B3:F3")
    t = cover.cell(row=3, column=2, value=title.upper())
    t.font = Font(name="Arial", bold=True, size=18, color=COLOR_HEADER_FONT)
    t.fill = _make_header_fill()
    t.alignment = Alignment(horizontal="center", vertical="center")
    cover.row_dimensions[3].height = 40

    cover.merge_cells("B5:F5")
    sub = cover.cell(row=5, column=2, value=f"Contains {len(sheets)} sheet(s): " + ", ".join(s.get("sheet_name", "") for s in sheets))
    sub.font = Font(name="Arial", size=11, color="444444")
    sub.alignment = Alignment(horizontal="center")

    cover.column_dimensions["A"].width = 3
    for col in ["B", "C", "D", "E", "F"]:
        cover.column_dimensions[col].width = 18

    # ── Data sheets ───────────────────────────────────────────────────────────
    for sheet_data in sheets:
        write_sheet(wb, sheet_data)

    # ── Save to BytesIO ───────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output